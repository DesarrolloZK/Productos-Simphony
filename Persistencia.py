#Libreria json: Contiene condigo para manejar los archivos JSON
import os
import json
import openpyxl as exc

#Esta clase nos permite manejar todos los archivos necesarios para el funcionamiento de la aplicacion
class Archivos():
    
    #Escribe el excel con la informacion de los productos
    def makeExcel(self,excel:str,hoja:str,cabeceras:str,datos:list)->bool:
        try:
            if os.path.exists(excel):
                arcExcel=exc.load_workbook(excel)
                while True:
                    if hoja in arcExcel.sheetnames:
                        excHoja=arcExcel[hoja]
                        excHoja.delete_rows(1,excHoja.max_row)
                        excHoja.append(cabeceras)
                        list(map(lambda x:excHoja.append(x),datos))
                        arcExcel.save(excel)
                        arcExcel.close()
                        print(f'{hoja} Creado')
                        return True
                    else:
                        excHoja=arcExcel.create_sheet(hoja)
                        arcExcel.save(excel)
                        arcExcel.close()
            else:
                print('Creando excel')
                arcExcel=exc.Workbook()
                excHoja=arcExcel.active
                excHoja.title=hoja
                arcExcel.save(excel)
                arcExcel.close()
                return self.makeExcel(excel,hoja,cabeceras,datos)
        except Exception as e:
            print(f'Error sobre el excel -> {e}')
            return False

    #Lee el archivo "Estaciones.json" y retorna un diccionario con dicha informacion
    def traerEstaciones()->dict:
        try:
            with open('Estaciones.json') as jf:
                configuracion=json.load(jf)
                jf.close()
                return configuracion
        except FileNotFoundError:
            return {}

    #Lee el archivo "configuracion.json" y retorna un diccionario con dicha informacion
    def traerConfiguraciones()->dict:
        try:
            with open('Config.json') as jf:
                configuracion=json.load(jf)
                jf.close()
                return configuracion
        except FileNotFoundError:
            return {}

if __name__=="__main__":
    print(Archivos.traerConfiguraciones())